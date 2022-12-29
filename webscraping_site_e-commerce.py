from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import *
from time import sleep
import openpyxl

class webscraping:

    def iniciar(self):
        
        self.raspagem_de_dados()
        self.criar_planilha()
        



    def raspagem_de_dados(self):
        self.nav = webdriver.Chrome()

        self.link = ('https://telefonesimportados.netlify.app/')
        print(self.nav.title)
        self.lista_nome_celulares = []
        self.lista_preco_celulares = []
        self.nav.get(self.link)
        sleep(2)
        for p in range(5):
            item = 1
            for i in range(12):
                lista_nomes = self.nav.find_element(By.XPATH,
                    f'/html/body/div[5]/div[2]/div[1]/div[{item}]/div/h2/a')  
                self.lista_nome_celulares.append(lista_nomes.text)
                sleep(1)
                lista_precos = self.nav.find_element(By.XPATH,
                    f'//div[{item}]/div[@class="single-shop-product" and 1]/div[@class="product-carousel-price" and 2]/ins[1]')
                self.lista_preco_celulares.append(lista_precos.text)
                item += 1
                sleep(1)
            try:
                botao_proximo = self.nav.find_element(By.XPATH,
                    '/html/body/div[5]/div[2]/div[2]/div/div/nav/ul/li[7]/a')
                botao_proximo.click()
                print(f'\u001b[32m{"Navegando para proxima pagina"}\u001b[0m')
                sleep(2)

            except NoSuchElementException:

                print(f'\u001b[33m{"Não há mais paginas!"}\u001b[0m')
                print(f'\u001b[32m{"Escaneamento Concluido"}\u001b[0m')
                self.nav.quit()

    def criar_planilha(self):
        index = 2
        planilha = openpyxl.Workbook()
        celulares = planilha['Sheet']
        celulares.title = 'Celulares'
        celulares['A1'] = 'Nome'
        celulares['B1'] = 'Preço'
        for nome, preco in zip(self.lista_nome_celulares, self.lista_preco_celulares):
            celulares.cell(column=1, row=index, value=nome)
            celulares.cell(column=2, row=index, value=preco)
            index += 1
        planilha.save("planilha_de_preços.xlsx")

        print(f'\u001b[32m{"Planilha criada com sucesso"}\u001b[0m')
start = webscraping()
start.iniciar()        
